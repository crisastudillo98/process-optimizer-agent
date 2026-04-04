from __future__ import annotations
import json
from pathlib import Path
from typing import Union

import fitz                      # PyMuPDF
import openpyxl
from langchain_core.documents import Document

from observability.logger import get_logger

logger = get_logger(__name__)


# ─────────────────────────────────────────────
# LOADERS POR TIPO
# ─────────────────────────────────────────────

def _load_pdf(path: Path) -> str:
    """Extrae texto de todas las páginas del PDF."""
    doc = fitz.open(str(path))
    pages = [page.get_text("text") for page in doc]
    doc.close()
    text = "\n\n".join(pages).strip()
    logger.info(f"PDF cargado: {path.name} — {len(pages)} páginas")
    return text


def _load_excel(path: Path) -> str:
    """
    Convierte cada hoja del Excel a texto tabular.
    Útil cuando el AS-IS viene en formato tabla de actividades.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            # Filtra filas completamente vacías
            cleaned = [str(cell) if cell is not None else "" for cell in row]
            if any(c.strip() for c in cleaned):
                rows.append(" | ".join(cleaned))
        if rows:
            sections.append(f"### Hoja: {sheet_name}\n" + "\n".join(rows))

    logger.info(f"Excel cargado: {path.name} — {len(wb.sheetnames)} hojas")
    return "\n\n".join(sections)


def _load_json(path: Path) -> str:
    """Serializa el JSON a texto legible para el LLM."""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    logger.info(f"JSON cargado: {path.name}")
    return json.dumps(data, ensure_ascii=False, indent=2)


def _load_text(path: Path) -> str:
    """Carga texto plano (.txt, .md)."""
    text = path.read_text(encoding="utf-8").strip()
    logger.info(f"Texto cargado: {path.name} — {len(text)} caracteres")
    return text


# ─────────────────────────────────────────────
# DISPATCHER PRINCIPAL
# ─────────────────────────────────────────────

LOADERS = {
    ".pdf":  _load_pdf,
    ".xlsx": _load_excel,
    ".xls":  _load_excel,
    ".json": _load_json,
    ".txt":  _load_text,
    ".md":   _load_text,
}


def load_document(source: Union[str, Path]) -> str:
    """
    Carga un documento desde ruta o texto libre.

    Args:
        source: Ruta al archivo (str/Path) o texto AS-IS directo.

    Returns:
        Texto extraído listo para el LLM.

    Raises:
        ValueError: Si la extensión no está soportada.
        FileNotFoundError: Si el archivo no existe.
    """
    # Si es texto libre (más de 260 chars o no tiene extensión válida), lo devuelve directo
    source_str = str(source)
    if isinstance(source, str) and (
        len(source_str) > 260 or Path(source_str).suffix.lower() not in LOADERS
    ):
        logger.info("Entrada detectada como texto libre")
        return source_str.strip()

    path = Path(source)

    if not path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {path}")

    suffix = path.suffix.lower()
    loader_fn = LOADERS.get(suffix)

    if loader_fn is None:
        raise ValueError(
            f"Extensión '{suffix}' no soportada. "
            f"Extensiones válidas: {list(LOADERS.keys())}"
        )

    return loader_fn(path)

def load_document_as_langchain(source: Union[str, Path]) -> list[Document]:
    """
    Versión LangChain-compatible que retorna lista de Documents.
    Útil para el pipeline RAG (chunking + embeddings).
    """
    text = load_document(source)
    source_str = str(source)
    return [Document(page_content=text, metadata={"source": source_str})]


# ─────────────────────────────────────────────
# NODO LANGGRAPH
# ─────────────────────────────────────────────

from models.schemas import AgentState   # import aquí para evitar circular

def node_load_document(state: AgentState) -> dict:
    """
    Nodo LangGraph: carga y parsea el documento de entrada.

    Entrada del estado: raw_input (texto libre) o input_file_path (ruta)
    Salida al estado:   raw_input (texto extraído), errors
    """
    logger.info("── Nodo: load_document ──")

    source = state.input_file_path or state.raw_input

    if not source:
        return {
            "errors": state.errors + ["No se proporcionó entrada (texto ni archivo)."],
            "current_node": "load_document",
        }

    try:
        text = load_document(source)

        if len(text.strip()) < 50:
            return {
                "errors": state.errors + [
                    "El documento cargado tiene muy poco contenido "
                    f"({len(text)} caracteres). Verifica el archivo."
                ],
                "current_node": "load_document",
            }

        logger.info(f"Documento cargado: {len(text)} caracteres")
        return {
            "raw_input":    text,
            "current_node": "load_document",
        }

    except (FileNotFoundError, ValueError) as e:
        logger.error(f"Error al cargar documento: {e}")
        return {
            "errors": state.errors + [f"load_document: {str(e)}"],
            "current_node": "load_document",
        }